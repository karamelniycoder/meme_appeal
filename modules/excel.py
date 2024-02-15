from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl import Workbook, load_workbook
from datetime import datetime
from loguru import logger
from time import sleep
import threading


class Excel:
    def __init__(self, total_len: int):
        workbook = Workbook()
        sheet = workbook.active
        self.lock = threading.Lock()
        self.file_name = f'{total_len}accs_{datetime.now().strftime("%d_%m_-%H_%M_%S")}.xlsx'

        sheet['A1'] = 'Index'
        sheet['B1'] = 'PrivateKey'
        sheet['C1'] = 'Address'
        sheet['D1'] = 'DiscordToken'
        sheet['E1'] = 'Proxy'
        sheet['F1'] = 'Status'

        for cell in sheet._cells:
            sheet.cell(cell[0], cell[1]).font = Font(bold=True)
            sheet.cell(cell[0], cell[1]).alignment = Alignment(horizontal='center')

        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 45
        sheet.column_dimensions['D'].width = 88
        sheet.column_dimensions['E'].width = 43
        sheet.column_dimensions['F'].width = 35

        workbook.save('results/'+self.file_name)


    def add_account(self, index: str, privatekey: str, address: str, token: str, proxy: str, status: str):
        with self.lock:
            while True:
                try:
                    workbook = load_workbook('results/'+self.file_name)
                    sheet = workbook.active

                    max_row = sheet.max_row + 1

                    valid_info = [
                        index,
                        privatekey,
                        address,
                        token,
                        proxy,
                        status,
                    ]
                    sheet.append(valid_info)

                    if '✅' in status: rgb_color = '32CD32'
                    else: rgb_color = 'ff0f0f'
                    sheet.cell(max_row, sheet.max_column).fill = PatternFill(patternType='solid', fgColor=Color(rgb=rgb_color))

                    workbook.save('results/'+self.file_name)
                    return True
                except PermissionError:
                    logger.warning(f'Excel | Cant save excel file, close it!')
                    sleep(3)
                except Exception as err:
                    logger.critical(f'Excel | Cant save excel file: {err} | {token[-20:]}')
                    return False